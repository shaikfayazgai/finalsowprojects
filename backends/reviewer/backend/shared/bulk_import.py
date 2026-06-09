"""
Reusable CSV/Excel bulk-import engine for onboarding users.

Used by universities (students), women-team (members), enterprise (employees),
and superadmin. Two-phase:
  - commit=False → parse + validate, return preview rows with per-row errors
                   and duplicate flags (NO DB writes).
  - commit=True  → insert selected rows, generate temp passwords, optionally
                   email credentials, write audit.

Supports both .csv and .xlsx (via openpyxl). Duplicate detection is against
both the in-file set and an `existing_emails` set the caller pre-fetches.
"""

from __future__ import annotations

import csv
import io
from typing import Any, Iterable

REQUIRED_COLUMNS = {"email"}


def _normalise_header(h: str | None) -> str:
    return (h or "").strip().lower().replace(" ", "")


def parse_rows(filename: str, raw: bytes) -> tuple[list[dict[str, str]], list[str]]:
    """Return (rows, headers). Each row is a dict of normalised-header → value.
    Raises ValueError for unsupported file types."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        return _parse_excel(raw)
    if name.endswith((".csv", ".txt")):
        return _parse_csv(raw)
    raise ValueError("Upload a .csv or .xlsx file.")


def _parse_csv(raw: bytes) -> tuple[list[dict[str, str]], list[str]]:
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")
    reader = csv.DictReader(io.StringIO(text))
    headers = [_normalise_header(h) for h in (reader.fieldnames or [])]
    rows: list[dict[str, str]] = []
    for raw_row in reader:
        rows.append({_normalise_header(k): (v or "").strip() for k, v in raw_row.items()})
    return rows, headers


def _parse_excel(raw: bytes) -> tuple[list[dict[str, str]], list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [_normalise_header(str(h) if h is not None else "") for h in header_row]
    rows: list[dict[str, str]] = []
    for values in rows_iter:
        if values is None or all(v is None or str(v).strip() == "" for v in values):
            continue
        row = {}
        for i, h in enumerate(headers):
            v = values[i] if i < len(values) else None
            row[h] = "" if v is None else str(v).strip()
        rows.append(row)
    return rows, headers


def validate_rows(
    rows: list[dict[str, str]],
    *,
    headers: list[str],
    existing_emails: set[str],
    allowed_roles: set[str] | None = None,
    required_extra: set[str] | None = None,
) -> dict[str, Any]:
    """Validate parsed rows. Returns {valid_rows, error_rows, headers}.
    valid_rows carry `_password` (may be empty) for the commit phase."""
    header_set = set(headers)
    missing_cols = (REQUIRED_COLUMNS | (required_extra or set())) - header_set
    if missing_cols:
        return {"missing_columns": sorted(missing_cols), "valid_rows": [], "error_rows": []}
    has_name = "name" in header_set or ("firstname" in header_set and "lastname" in header_set)

    valid_rows: list[dict[str, Any]] = []
    error_rows: list[dict[str, Any]] = []
    seen_in_file: set[str] = set()

    for idx, row in enumerate(rows, start=2):  # row 1 = header
        email = row.get("email", "").lower()
        first = row.get("firstname", "")
        last = row.get("lastname", "")
        name = row.get("name", "") or f"{first} {last}".strip()
        role = (row.get("role", "") or "").lower()
        dept = row.get("department") or None

        errors: list[str] = []
        is_duplicate = False
        if not email or "@" not in email:
            errors.append("Invalid email")
        if has_name and not name:
            errors.append("Missing name")
        if allowed_roles and role and role not in allowed_roles:
            errors.append(f"Role must be one of {sorted(allowed_roles)}")
        if email and email in seen_in_file:
            errors.append("Duplicate email in file")
            is_duplicate = True
        elif email and email in existing_emails:
            errors.append("Email already exists in system")
            is_duplicate = True

        record = {
            "rowNumber": idx,
            "email": email,
            "name": name,
            "firstName": first or (name.split(" ")[0] if name else ""),
            "lastName": last or (" ".join(name.split(" ")[1:]) if name and " " in name else ""),
            "role": role,
            "department": dept,
            "phone": row.get("phone") or None,
            "isDuplicate": is_duplicate,
            "selectable": not is_duplicate,  # frontend pre-checks non-duplicates
        }
        if errors:
            error_rows.append({**record, "errors": errors})
        else:
            seen_in_file.add(email)
            valid_rows.append({**record, "_password": row.get("password") or ""})

    return {"valid_rows": valid_rows, "error_rows": error_rows, "missing_columns": []}


def import_summary(valid_rows: Iterable[dict], error_rows: list[dict],
                   *, committed: bool, inserted: int,
                   email_results: list[dict] | None = None) -> dict[str, Any]:
    email_results = email_results or []
    valid_list = list(valid_rows)
    return {
        "totalRows": len(valid_list) + len(error_rows),
        "validRows": [{k: v for k, v in r.items() if not k.startswith("_")} for r in valid_list],
        "errorRows": error_rows,
        "committed": committed,
        "insertedCount": inserted,
        "emailSentCount": sum(1 for r in email_results if r.get("sent")),
        "emailFailedCount": sum(1 for r in email_results if not r.get("sent")),
        "emailResults": email_results,
    }
