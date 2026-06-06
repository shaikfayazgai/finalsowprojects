#!/usr/bin/env python3
"""GlimmoraTeam — standard MNC-style feature / delivery tracker (master register)."""

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink

OUTPUT = Path(__file__).resolve().parents[1] / "docs" / "delivery" / "GlimmoraTeam-Feature-Tracker.xlsx"
TODAY = date(2026, 5, 28)

# ── Dropdown values (standard PMO vocabulary) ──
OVERALL_STATUS = ["Not Started", "In Progress", "On Hold", "Completed", "Cancelled", "Deferred"]
FE_BE_STATUS = ["Not Started", "In Progress", "Complete", "Blocked", "N/A"]
PRIORITY = ["P1 — Critical", "P2 — High", "P3 — Medium", "P4 — Low"]
PHASE = ["Phase 1 — MVP", "Phase 1 — Hardening", "Phase 2", "Backlog"]
RAG = ["Green", "Amber", "Red", "Grey"]
ASSIGNEES = (
    "Unassigned,Product Manager,Engineering Manager,Frontend Lead,Backend Lead,"
    "Full Stack Developer,QA Lead,DevOps / SRE,Security Lead,UI/UX Designer,"
    "Integrations Lead,AI / ML Engineer"
)

# id, module, feature, description, priority, phase, overall, fe, be, assignee, start, due, dependency, remarks
_ROWS = [
    ("FTR-001", "Platform", "Application shell & routing", "Multi-portal Next.js App Router with role-based layouts", "P1 — Critical", "Phase 1 — MVP", "Completed", "Complete", "Complete", "Frontend Lead", "2026-03-01", "2026-04-30", "—", "Production-ready"),
    ("FTR-002", "Platform", "Authentication (email / password)", "Login, registration, password reset, session management", "P1 — Critical", "Phase 1 — MVP", "Completed", "Complete", "Complete", "Backend Lead", "2026-03-01", "2026-04-15", "FTR-008", "NextAuth v5 + Prisma"),
    ("FTR-003", "Platform", "OAuth (Google & Microsoft)", "SSO sign-in for contributors and enterprise users", "P1 — Critical", "Phase 1 — MVP", "In Progress", "Complete", "In Progress", "Backend Lead", "2026-04-01", "2026-06-15", "FTR-002", "Enterprise IdP routing partial"),
    ("FTR-004", "Platform", "MFA / TOTP", "Multi-factor authentication setup and verification", "P2 — High", "Phase 1 — Hardening", "In Progress", "Complete", "In Progress", "Security Lead", "2026-05-01", "2026-06-30", "FTR-002", "E2E coverage pending"),
    ("FTR-005", "Platform", "OTP (email & phone)", "One-time password verification flows", "P2 — High", "Phase 1 — MVP", "Completed", "Complete", "Complete", "Backend Lead", "2026-04-01", "2026-05-01", "FTR-011", "Live"),
    ("FTR-006", "Platform", "Role-based route guards", "Middleware and hooks enforcing portal access", "P1 — Critical", "Phase 1 — MVP", "In Progress", "Complete", "In Progress", "Platform Lead", "2026-04-15", "2026-06-01", "FTR-002", "Demo bypass flags in dev"),
    ("FTR-007", "Platform", "API authorization (RBAC audit)", "requireRole on all API route handlers", "P1 — Critical", "Phase 1 — Hardening", "In Progress", "N/A", "In Progress", "Backend Lead", "2026-05-15", "2026-06-15", "FTR-006", "2 of 28 routes wired"),
    ("FTR-008", "Platform", "Database schema (Prisma + Postgres)", "Core domain models and migrations", "P1 — Critical", "Phase 1 — MVP", "In Progress", "N/A", "In Progress", "Backend Lead", "2026-03-15", "2026-06-30", "—", "Schema exists; many domains mock-backed"),
    ("FTR-009", "Platform", "Observability (Sentry / logging)", "Error tracking, structured logs, alerting", "P1 — Critical", "Phase 1 — Hardening", "Not Started", "N/A", "Not Started", "DevOps / SRE", "2026-06-01", "2026-07-15", "—", "Production blocker"),
    ("FTR-010", "Platform", "Security headers & CSRF", "CSP, rate limiting, CSRF middleware", "P1 — Critical", "Phase 1 — Hardening", "Not Started", "N/A", "Not Started", "Security Lead", "2026-06-15", "2026-07-01", "FTR-007", ""),
    ("FTR-011", "Platform", "Transactional email", "Nodemailer + React Email templates", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "In Progress", "Backend Lead", "2026-04-01", "2026-06-01", "—", "Rate limit + RBAC audit pending"),
    ("FTR-012", "Platform", "Razorpay — order creation", "Enterprise milestone payment order intent", "P1 — Critical", "Phase 1 — MVP", "Completed", "Complete", "Complete", "Backend Lead", "2026-04-01", "2026-05-15", "FTR-048", "Gated by acceptance decision"),
    ("FTR-013", "Platform", "Razorpay — webhooks", "HMAC signature verification, idempotent events", "P1 — Critical", "Phase 1 — MVP", "Completed", "N/A", "Complete", "Backend Lead", "2026-04-15", "2026-05-15", "FTR-012", "Shipped"),
    ("FTR-014", "Platform", "Razorpay — payouts", "Platform to contributor payout leg", "P2 — High", "Phase 1 — Hardening", "Not Started", "In Progress", "Not Started", "Backend Lead", "2026-07-01", "2026-07-30", "FTR-012", "Conditional on payout go-live"),
    ("FTR-015", "Platform", "Audit log", "Immutable audit trail for state changes", "P1 — Critical", "Phase 1 — MVP", "In Progress", "Complete", "In Progress", "Backend Lead", "2026-04-01", "2026-06-15", "FTR-008", "Acceptance decisions persisted"),
    ("FTR-016", "Platform", "Subscription & plan gating", "SaaS plan limits, usage, feature gates", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "In Progress", "Full Stack Developer", "2026-05-01", "2026-05-28", "FTR-008", "Sidebar plan chip shipped"),
    ("FTR-017", "Platform", "Backend API handoff", "External backend team scaffold + contract", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "In Progress", "Engineering Manager", "2026-05-01", "2026-05-28", "—", "Frontend owns Prisma"),
    ("FTR-020", "Contributor", "Dashboard", "Contributor home — tasks, progress, alerts", "P1 — Critical", "Phase 1 — MVP", "In Progress", "Complete", "In Progress", "Frontend Lead", "2026-04-01", "2026-05-01", "FTR-021", "Mock / seed data"),
    ("FTR-021", "Contributor", "Task workroom & lifecycle", "Assigned → submit → review → approved flow", "P1 — Critical", "Phase 1 — MVP", "In Progress", "Complete", "In Progress", "Full Stack Developer", "2026-04-15", "2026-06-30", "FTR-090", "localStorage today"),
    ("FTR-022", "Contributor", "Submissions & revisions", "Deliverable submit and rework loop", "P1 — Critical", "Phase 1 — MVP", "In Progress", "Complete", "In Progress", "Full Stack Developer", "2026-04-15", "2026-06-30", "FTR-021", ""),
    ("FTR-023", "Contributor", "Digital twin profile", "Skills, evidence, portfolio", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "In Progress", "Full Stack Developer", "2026-04-01", "2026-06-15", "—", ""),
    ("FTR-024", "Contributor", "Onboarding flow", "Multi-step contributor registration", "P1 — Critical", "Phase 1 — MVP", "In Progress", "Complete", "In Progress", "Full Stack Developer", "2026-04-01", "2026-06-01", "FTR-002", ""),
    ("FTR-025", "Contributor", "Earnings & payouts view", "Contributor earnings dashboard", "P2 — High", "Phase 1 — Hardening", "In Progress", "Complete", "Not Started", "Frontend Lead", "2026-06-01", "2026-07-15", "FTR-014", "Mock data"),
    ("FTR-026", "Contributor", "Credentials wallet", "Issue, share, verify credentials", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "In Progress", "Full Stack Developer", "2026-05-01", "2026-06-30", "—", "Public share page works"),
    ("FTR-027", "Contributor", "Support & tickets", "FAQs, grievances, safety reports", "P3 — Medium", "Phase 1 — MVP", "In Progress", "Complete", "In Progress", "Full Stack Developer", "2026-05-01", "2026-07-01", "—", ""),
    ("FTR-028", "Contributor", "Community & messages", "Community feed and messaging", "P3 — Medium", "Phase 2", "Deferred", "Complete", "Not Started", "Frontend Lead", "2026-07-01", "2026-08-01", "—", "Visual-only mock"),
    ("FTR-029", "Contributor", "Learning recommendations", "AI skill learning suggestions", "P3 — Medium", "Phase 2", "Deferred", "Complete", "Not Started", "Frontend Lead", "2026-07-01", "2026-08-01", "FTR-097", ""),
    ("FTR-030", "Contributor", "Account auth settings", "Password / OAuth / SSO mode selection", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "In Progress", "Full Stack Developer", "2026-05-01", "2026-05-28", "FTR-003", "Recently shipped"),
    ("FTR-040", "Enterprise", "Dashboard", "Executive dashboard with KPIs and plan usage", "P1 — Critical", "Phase 1 — MVP", "In Progress", "Complete", "In Progress", "Frontend Lead", "2026-04-01", "2026-05-01", "FTR-016", ""),
    ("FTR-041", "Enterprise", "SOW repository", "SOW list and detail views", "P1 — Critical", "Phase 1 — MVP", "In Progress", "Complete", "In Progress", "Full Stack Developer", "2026-04-01", "2026-06-15", "FTR-091", "API + mocks hybrid"),
    ("FTR-042", "Enterprise", "SOW intake (upload + AI)", "5-stage intake wizard with AI extraction", "P1 — Critical", "Phase 1 — MVP", "In Progress", "Complete", "In Progress", "Full Stack Developer", "2026-04-15", "2026-06-30", "FTR-097", "Commit local-state"),
    ("FTR-043", "Enterprise", "SOW approval pipeline", "5-stage commercial / legal sign-off", "P1 — Critical", "Phase 1 — MVP", "In Progress", "Complete", "In Progress", "Full Stack Developer", "2026-04-15", "2026-06-30", "FTR-041", ""),
    ("FTR-044", "Enterprise", "Decomposition workspace", "Break SOW into workstreams and tasks", "P1 — Critical", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Full Stack Developer", "2026-05-01", "2026-07-01", "FTR-091", "Mock plans"),
    ("FTR-045", "Enterprise", "Team matching & formation", "Suggest and assign contributors", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Frontend Lead", "2026-05-15", "2026-07-15", "FTR-044", ""),
    ("FTR-046", "Enterprise", "Projects portfolio", "Active and completed project tiles", "P1 — Critical", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Full Stack Developer", "2026-05-01", "2026-07-01", "FTR-091", ""),
    ("FTR-047", "Enterprise", "Delivery tracking", "Milestones, exceptions, lifecycle view", "P1 — Critical", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Full Stack Developer", "2026-05-01", "2026-07-01", "FTR-021", ""),
    ("FTR-048", "Enterprise", "Review & acceptance queue", "Accept / rework deliverables", "P1 — Critical", "Phase 1 — MVP", "In Progress", "Complete", "In Progress", "Full Stack Developer", "2026-04-15", "2026-06-01", "FTR-015", "Postgres audit log"),
    ("FTR-049", "Enterprise", "Reviewer sub-portal", "QA inbox, task monitor, metrics", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Frontend Lead", "2026-05-01", "2026-07-01", "—", ""),
    ("FTR-050", "Enterprise", "Billing overview", "Budget, invoices, rate cards UI", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Full Stack Developer", "2026-05-01", "2026-07-15", "FTR-012", ""),
    ("FTR-051", "Enterprise", "Workforce roster", "Internal employee list and profile drawer", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Full Stack Developer", "2026-05-15", "2026-05-28", "FTR-094", "Mock localStorage"),
    ("FTR-052", "Enterprise", "Workforce CSV import", "Upload, preview, apply roster", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Full Stack Developer", "2026-05-15", "2026-06-30", "FTR-094", ""),
    ("FTR-053", "Enterprise", "Manual add employee", "Form-based roster provisioning", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Full Stack Developer", "2026-05-15", "2026-05-28", "FTR-051", ""),
    ("FTR-054", "Enterprise", "Settings — tenant & plan", "RBAC settings navigation and workspaces", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Full Stack Developer", "2026-05-01", "2026-06-15", "FTR-093", ""),
    ("FTR-055", "Enterprise", "Profile & security settings", "Sessions, MFA display, security prefs", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "In Progress", "Full Stack Developer", "2026-05-01", "2026-06-15", "FTR-004", ""),
    ("FTR-056", "Enterprise", "Notifications centre", "In-app notification feed", "P3 — Medium", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Frontend Lead", "2026-05-15", "2026-07-01", "FTR-093", ""),
    ("FTR-057", "Enterprise", "Enterprise onboarding", "First-time admin onboarding flow", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "In Progress", "Full Stack Developer", "2026-04-15", "2026-06-01", "—", ""),
    ("FTR-058", "Enterprise", "Analytics console", "Economic and governance analytics", "P4 — Low", "Phase 2", "Deferred", "Complete", "Not Started", "Unassigned", "—", "2026-09-01", "—", "Phase 2 scope"),
    ("FTR-059", "Enterprise", "Compliance module", "ESG, evidence locker, PODL", "P4 — Low", "Phase 2", "Deferred", "Complete", "Not Started", "Unassigned", "—", "2026-09-01", "—", "Phase 2 scope"),
    ("FTR-060", "Enterprise", "Audit log export", "Enterprise audit trail export", "P4 — Low", "Phase 2", "Deferred", "In Progress", "In Progress", "Backend Lead", "2026-07-01", "2026-08-01", "FTR-015", ""),
    ("FTR-070", "Mentor", "Dashboard", "Mentor home with queue summary", "P1 — Critical", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Frontend Lead", "2026-04-01", "2026-05-01", "—", ""),
    ("FTR-071", "Mentor", "Review queue", "Pending submissions awaiting review", "P1 — Critical", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Full Stack Developer", "2026-04-15", "2026-06-30", "FTR-092", ""),
    ("FTR-072", "Mentor", "Active review detail", "Rubric scoring and approve / rework", "P1 — Critical", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Full Stack Developer", "2026-04-15", "2026-06-30", "FTR-092", ""),
    ("FTR-073", "Mentor", "Escalation management", "Escalated review resolution", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Frontend Lead", "2026-05-01", "2026-07-01", "—", ""),
    ("FTR-074", "Mentor", "Mentorship sessions", "Session notes and history", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Frontend Lead", "2026-05-15", "2026-07-15", "—", ""),
    ("FTR-075", "Mentor", "Review history", "Completed reviews archive", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Frontend Lead", "2026-05-01", "2026-06-01", "—", ""),
    ("FTR-076", "Mentor", "Profile & settings", "Mentor preferences", "P3 — Medium", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Frontend Lead", "2026-05-15", "2026-07-01", "—", ""),
    ("FTR-080", "Admin", "Dashboard", "Platform operations overview", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Frontend Lead", "2026-04-01", "2026-06-01", "—", ""),
    ("FTR-081", "Admin", "Tenant provisioning wizard", "Create tenant, tier, invite admin", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Full Stack Developer", "2026-05-01", "2026-07-01", "FTR-008", ""),
    ("FTR-082", "Admin", "Tenant subscription management", "Plan, usage, change history", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "In Progress", "Full Stack Developer", "2026-05-01", "2026-06-15", "FTR-016", ""),
    ("FTR-083", "Admin", "User management", "Roles, invitations, directory", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Full Stack Developer", "2026-05-01", "2026-07-01", "FTR-007", ""),
    ("FTR-084", "Admin", "Email template editor", "SOW, OTP, invite templates", "P3 — Medium", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Frontend Lead", "2026-05-15", "2026-07-01", "FTR-011", ""),
    ("FTR-085", "Admin", "Platform pricing config", "Contributor rate tables by segment", "P3 — Medium", "Phase 1 — MVP", "In Progress", "Complete", "In Progress", "Full Stack Developer", "2026-05-01", "2026-06-15", "—", ""),
    ("FTR-086", "Admin", "SOW oversight", "Cross-tenant commercial sign-off", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Frontend Lead", "2026-05-01", "2026-07-01", "FTR-041", ""),
    ("FTR-087", "Admin", "KYC review queue", "Contributor onboarding case review", "P2 — High", "Phase 1 — MVP", "In Progress", "Complete", "Not Started", "Full Stack Developer", "2026-05-15", "2026-07-01", "FTR-024", ""),
    ("FTR-088", "Admin", "Governance & safety reports", "Platform-wide triage", "P4 — Low", "Phase 2", "Deferred", "In Progress", "Not Started", "Unassigned", "—", "2026-09-01", "—", ""),
    ("FTR-090", "Backend", "Task persistence (Postgres)", "Migrate task store from localStorage", "P1 — Critical", "Phase 1 — Hardening", "Not Started", "N/A", "Not Started", "Backend Lead", "2026-06-01", "2026-07-15", "FTR-008", "Slice 5 — prod blocker"),
    ("FTR-091", "Backend", "SOW / project / decomposition APIs", "Prisma models and CRUD endpoints", "P1 — Critical", "Phase 1 — Hardening", "In Progress", "N/A", "In Progress", "Backend Lead", "2026-05-15", "2026-07-30", "FTR-008", ""),
    ("FTR-092", "Backend", "Mentor decision persistence", "Persist approve / rework to database", "P1 — Critical", "Phase 1 — Hardening", "Not Started", "N/A", "Not Started", "Backend Lead", "2026-06-01", "2026-07-15", "FTR-090", ""),
    ("FTR-093", "Backend", "Settings & preferences API", "OperatorPreferences model + endpoints", "P2 — High", "Phase 1 — Hardening", "Not Started", "N/A", "Not Started", "Backend Lead", "2026-06-15", "2026-07-01", "FTR-008", ""),
    ("FTR-094", "Backend", "Workforce API (production)", "Wire workforce routes to Prisma", "P2 — High", "Phase 1 — Hardening", "In Progress", "N/A", "In Progress", "Backend Lead", "2026-05-15", "2026-07-01", "FTR-051", ""),
    ("FTR-095", "Backend", "Enterprise SSO / HRIS", "SAML, OIDC; CSV import path", "P2 — High", "Phase 1 — Hardening", "Not Started", "In Progress", "Not Started", "Integrations Lead", "2026-06-01", "2026-08-01", "FTR-003", ""),
    ("FTR-096", "Backend", "Outbound webhooks", "Jira, Slack, generic connectors", "P3 — Medium", "Phase 2", "Not Started", "N/A", "Not Started", "Integrations Lead", "2026-07-01", "2026-08-15", "—", ""),
    ("FTR-097", "Backend", "Assistive AI agents", "SOW intake, decomposition, review AI", "P2 — High", "Phase 1 — Hardening", "In Progress", "In Progress", "In Progress", "AI / ML Engineer", "2026-05-01", "2026-08-01", "—", ""),
    ("FTR-098", "Backend", "Server-side exports", "PDF / CSV report generation", "P3 — Medium", "Phase 2", "Not Started", "N/A", "Not Started", "Backend Lead", "2026-07-01", "2026-08-01", "—", ""),
    ("FTR-099", "Backend", "Real-time updates (SSE / WS)", "Live activity without refresh", "P3 — Medium", "Phase 2", "Not Started", "N/A", "Not Started", "DevOps / SRE", "2026-08-01", "2026-09-01", "—", ""),
    ("FTR-100", "Backend", "WCAG 2.1 AA audit", "Accessibility compliance gate", "P2 — High", "Phase 1 — Hardening", "Not Started", "In Progress", "Not Started", "QA Lead", "2026-07-01", "2026-08-30", "—", "Phase 1D gate"),
]

HEADERS = [
    "Feature ID",
    "Module",
    "Feature Name",
    "Description",
    "Priority",
    "Phase",
    "Overall Status",
    "Frontend Status",
    "Backend Status",
    "% Complete",
    "RAG",
    "Assigned To",
    "Accountable (EM)",
    "Planned Start",
    "Planned End",
    "Actual End",
    "Last Updated",
    "Dependency ID",
    "Remarks",
]

WIDTHS = [10, 11, 26, 34, 12, 16, 13, 13, 13, 10, 8, 16, 16, 12, 12, 12, 12, 12, 28]

DATA_ROW = 12  # first data row

# Styling
NAVY = "1F3864"
HDR_FILL = PatternFill("solid", fgColor=NAVY)
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PCT_FORMULA = (
    '=IF(G{row}="Completed",100,IF(G{row}="Deferred",0,IF(G{row}="Cancelled",0,'
    'IF(G{row}="Not Started",0,IF(G{row}="On Hold",'
    'IF(OR(H{row}="Complete",H{row}="N/A"),50,25),'
    'ROUND((IF(H{row}="Complete",100,IF(H{row}="In Progress",50,IF(H{row}="Blocked",25,0)))+'
    'IF(I{row}="Complete",100,IF(I{row}="In Progress",50,IF(I{row}="Blocked",25,IF(I{row}="N/A",100,0)))))/2,0))))))'
)

RAG_FORMULA = (
    '=IF(G{row}="Completed","Green",IF(G{row}="Deferred","Grey",IF(G{row}="Cancelled","Grey",'
    'IF(G{row}="On Hold","Red",IF(J{row}>=80,"Green",IF(J{row}>=40,"Amber","Red"))))))'
)


def _font(**kw):
    return Font(name="Calibri", **kw)


def _cell(ws, r, c, val, *, bold=False, size=10, color="000000", fill=None, align="left", wrap=False, fmt=None, italic=False):
    x = ws.cell(r, c, val)
    x.font = _font(bold=bold, size=size, color=color, italic=italic)
    x.border = BD
    x.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill:
        x.fill = fill
    if fmt:
        x.number_format = fmt
    return x


def _parse_date(s):
    if not s or s == "—":
        return None
    return date.fromisoformat(s)


def _project_cover(wb):
    ws = wb.active
    ws.title = "Project Info"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    _cell(ws, 1, 1, "GlimmoraTeam — Feature Delivery Tracker", bold=True, size=18, color=NAVY)
    ws.merge_cells("A2:D2")
    _cell(ws, 2, 1, "Standard program register · Phase 1 MVP", size=11, color="595959")

    fields = [
        ("Project name", "GlimmoraTeam Global Workforce Platform"),
        ("Document owner", "Product / Engineering"),
        ("Version", "1.0"),
        ("Last updated", TODAY),
        ("Reporting period", "Phase 1 (Apr – Aug 2026)"),
        ("Project manager", "[Name]"),
        ("Engineering manager", "[Name]"),
        ("Status cadence", "Weekly — update Assigned To, statuses, and % Complete"),
    ]
    for i, (k, v) in enumerate(fields, 4):
        _cell(ws, i, 1, k, bold=True, color="595959")
        x = _cell(ws, i, 2, v)
        if k == "Last updated":
            x.number_format = "DD MMM YYYY"

    ws.merge_cells("A13:D13")
    _cell(ws, 13, 1, "Navigation", bold=True, size=12, color=NAVY)
    _cell(ws, 14, 1, "Sheet", bold=True, fill=HDR_FILL, color="FFFFFF", align="center")
    _cell(ws, 14, 2, "Purpose", bold=True, fill=HDR_FILL, color="FFFFFF", align="center")
    nav = [
        ("Feature Register", "Master tracker — edit here"),
        ("Dashboard", "Executive summary (auto-calculated)"),
        ("Resource Plan", "Assignee workload view"),
        ("Definitions", "Column and status definitions"),
    ]
    for i, (sheet, purpose) in enumerate(nav, 15):
        c = ws.cell(i, 1, sheet)
        c.hyperlink = Hyperlink(ref=c.coordinate, location=f"'{sheet}'!A1", display=sheet)
        c.font = _font(color="0563C1", underline="single")
        c.border = BD
        _cell(ws, i, 2, purpose)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 42


def _feature_register(wb):
    ws = wb.create_sheet("Feature Register")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:S1")
    _cell(ws, 1, 1, "Feature Register", bold=True, size=16, color=NAVY)
    ws.merge_cells("A2:S2")
    _cell(ws, 2, 1, "Single source of truth · Filter by Module, Assigned To, or Overall Status", size=10, color="595959", italic=True)

    hdr = 11
    for i, h in enumerate(HEADERS, 1):
        _cell(ws, hdr, i, h, bold=True, size=10, color="FFFFFF", fill=HDR_FILL, align="center", wrap=True)
    ws.row_dimensions[hdr].height = 36
    ws.freeze_panes = f"A{DATA_ROW}"
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    last = DATA_ROW + len(_ROWS) - 1
    for idx, row in enumerate(_ROWS):
        r = DATA_ROW + idx
        fid, mod, name, desc, pri, phase, overall, fe, be, assignee, start, due, dep, remarks = row
        vals = [
            fid, mod, name, desc, pri, phase, overall, fe, be,
            None, None, assignee, "Engineering Manager",
            _parse_date(start), _parse_date(due), None, TODAY, dep, remarks,
        ]
        for c, v in enumerate(vals, 1):
            if c == 10:
                ws.cell(r, c, PCT_FORMULA.format(row=r))
                ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(r, c).border = BD
            elif c == 11:
                ws.cell(r, c, RAG_FORMULA.format(row=r))
                ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(r, c).border = BD
            elif c == 3:
                _cell(ws, r, c, v, bold=True, wrap=True)
            elif c in (14, 15, 16, 17):
                _cell(ws, r, c, v, align="center", fmt="DD MMM YYYY")
            elif c in (5, 6, 7, 8, 9, 12, 13, 18):
                _cell(ws, r, c, v, align="center")
            else:
                _cell(ws, r, c, v, wrap=(c in (4, 19)))
        if idx % 2:
            for c in range(1, 20):
                if ws.cell(r, c).fill.fill_type is None:
                    ws.cell(r, c).fill = ALT_FILL

    ws.auto_filter.ref = f"A{hdr}:S{last}"

    # Validations
    def dv_list(opts, cols):
        d = DataValidation(type="list", formula1=f'"{",".join(opts)}"', allow_blank=True)
        ws.add_data_validation(d)
        for col in cols:
            d.add(f"{col}{DATA_ROW}:{col}{last}")

    dv_list(OVERALL_STATUS, ["G"])
    dv_list(FE_BE_STATUS, ["H", "I"])
    dv_list(PRIORITY, ["E"])
    dv_list(PHASE, ["F"])
    dv_list(RAG, ["K"])
    dv_list(ASSIGNEES.split(","), ["L", "M"])

    ws.conditional_formatting.add(
        f"J{DATA_ROW}:J{last}",
        ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                       mid_type="num", mid_value=50, mid_color="FFEB84",
                       end_type="num", end_value=100, end_color="63BE7B"),
    )
    for rag, color in [("Green", "C6EFCE"), ("Amber", "FFEB9C"), ("Red", "FFC7CE"), ("Grey", "D9D9D9")]:
        ws.conditional_formatting.add(
            f"K{DATA_ROW}:K{last}",
            FormulaRule(formula=[f'$K{DATA_ROW}="{rag}"'], fill=PatternFill("solid", fgColor=color)),
        )

    return ws, last


def _dashboard(wb, last):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    dr, dl = DATA_ROW, last

    ws.merge_cells("A1:F1")
    _cell(ws, 1, 1, "Program Dashboard", bold=True, size=16, color=NAVY)
    ws.merge_cells("A2:F2")
    _cell(ws, 2, 1, f"As of {TODAY:%d %b %Y} · sourced from Feature Register", size=10, color="595959", italic=True)

    kpis = [
        ("Total features", f"=COUNTA('Feature Register'!A{dr}:A{dl})"),
        ("Completed", f"=COUNTIF('Feature Register'!G{dr}:G{dl},\"Completed\")"),
        ("In progress", f"=COUNTIF('Feature Register'!G{dr}:G{dl},\"In Progress\")"),
        ("Not started", f"=COUNTIF('Feature Register'!G{dr}:G{dl},\"Not Started\")"),
        ("P1 critical (open)", f"=COUNTIFS('Feature Register'!E{dr}:E{dl},\"P1*\",'Feature Register'!G{dr}:G{dl},\"<>Completed\",'Feature Register'!G{dr}:G{dl},\"<>Deferred\",'Feature Register'!G{dr}:G{dl},\"<>Cancelled\")"),
        ("Avg % complete", f"=IFERROR(AVERAGE('Feature Register'!J{dr}:J{dl}),0)"),
        ("Red / Amber (RAG)", f"=COUNTIF('Feature Register'!K{dr}:K{dl},\"Red\")&\" / \"&COUNTIF('Feature Register'!K{dr}:K{dl},\"Amber\")"),
        ("Deferred (Phase 2)", f"=COUNTIF('Feature Register'!G{dr}:G{dl},\"Deferred\")"),
    ]
    for i, (label, formula) in enumerate(kpis):
        col = 1 + (i % 4) * 2
        row = 4 + (i // 4) * 3
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        _cell(ws, row, col, label, bold=True, size=9, color="595959", align="center")
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        c = _cell(ws, row + 1, col, formula, bold=True, size=18, color=NAVY, align="center")
        if "Avg" in label:
            c.number_format = "0.0"

    _cell(ws, 12, 1, "By module", bold=True, size=12, color=NAVY)
    mod_hdr = ["Module", "Total", "Completed", "In Progress", "Not Started", "Avg %"]
    for i, h in enumerate(mod_hdr, 1):
        _cell(ws, 13, i, h, bold=True, color="FFFFFF", fill=HDR_FILL, align="center")

    modules = ["Platform", "Contributor", "Enterprise", "Mentor", "Admin", "Backend"]
    for i, mod in enumerate(modules, 14):
        q = f"'Feature Register'!B{dr}:B{dl}"
        _cell(ws, i, 1, mod, bold=True)
        _cell(ws, i, 2, f'=COUNTIF({q},"{mod}")', align="center")
        _cell(ws, i, 3, f"=COUNTIFS('Feature Register'!B{dr}:B{dl},\"{mod}\",'Feature Register'!G{dr}:G{dl},\"Completed\")", align="center")
        _cell(ws, i, 4, f"=COUNTIFS('Feature Register'!B{dr}:B{dl},\"{mod}\",'Feature Register'!G{dr}:G{dl},\"In Progress\")", align="center")
        _cell(ws, i, 5, f"=COUNTIFS('Feature Register'!B{dr}:B{dl},\"{mod}\",'Feature Register'!G{dr}:G{dl},\"Not Started\")", align="center")
        _cell(ws, i, 6, f"=IFERROR(AVERAGEIF('Feature Register'!B{dr}:B{dl},\"{mod}\",'Feature Register'!J{dr}:J{dl}),0)", align="center")
        ws.cell(i, 6).number_format = "0.0"

    for col, w in zip("ABCDEF", [14, 8, 10, 10, 10, 10]):
        ws.column_dimensions[col].width = w


def _resource_plan(wb, last):
    ws = wb.create_sheet("Resource Plan")
    ws.sheet_view.showGridLines = False
    dr, dl = DATA_ROW, last

    ws.merge_cells("A1:F1")
    _cell(ws, 1, 1, "Resource Plan — by assignee", bold=True, size=16, color=NAVY)
    ws.merge_cells("A2:F2")
    _cell(ws, 2, 1, "Standard workload view · who owns how many open items", size=10, color="595959", italic=True)

    hdr = ["Assigned To", "Total assigned", "Completed", "In progress", "Not started", "Open P1 items"]
    for i, h in enumerate(hdr, 1):
        _cell(ws, 4, i, h, bold=True, color="FFFFFF", fill=HDR_FILL, align="center")

    assignees = [a for a in ASSIGNEES.split(",") if a != "Unassigned"]
    for i, name in enumerate(assignees, 5):
        _cell(ws, i, 1, name, bold=True)
        _cell(ws, i, 2, f"=COUNTIF('Feature Register'!L{dr}:L{dl},A{i})", align="center")
        _cell(ws, i, 3, f"=COUNTIFS('Feature Register'!L{dr}:L{dl},A{i},'Feature Register'!G{dr}:G{dl},\"Completed\")", align="center")
        _cell(ws, i, 4, f"=COUNTIFS('Feature Register'!L{dr}:L{dl},A{i},'Feature Register'!G{dr}:G{dl},\"In Progress\")", align="center")
        _cell(ws, i, 5, f"=COUNTIFS('Feature Register'!L{dr}:L{dl},A{i},'Feature Register'!G{dr}:G{dl},\"Not Started\")", align="center")
        _cell(ws, i, 6, f"=COUNTIFS('Feature Register'!L{dr}:L{dl},A{i},'Feature Register'!E{dr}:E{dl},\"P1*\",'Feature Register'!G{dr}:G{dl},\"<>Completed\")", align="center")

    for col, w in zip("ABCDEF", [22, 12, 12, 12, 12, 12]):
        ws.column_dimensions[col].width = w


def _definitions(wb):
    ws = wb.create_sheet("Definitions")
    ws.sheet_view.showGridLines = False
    _cell(ws, 1, 1, "Column & status definitions", bold=True, size=14, color=NAVY)

    rows = [
        ("Feature ID", "Unique reference (FTR-xxx). Used in dependencies."),
        ("Module", "Product area: Platform, Contributor, Enterprise, Mentor, Admin, Backend."),
        ("Overall Status", "Not Started → In Progress → Completed. Also: On Hold, Cancelled, Deferred."),
        ("Frontend / Backend Status", "Delivery layer status. Use N/A when not applicable (e.g. backend-only)."),
        ("% Complete", "Auto-calculated from overall + FE/BE status."),
        ("RAG", "Auto: Green (on track), Amber (attention), Red (blocked/overdue risk), Grey (deferred)."),
        ("Assigned To", "Person doing the work (developer / lead)."),
        ("Accountable (EM)", "Manager accountable for delivery sign-off."),
        ("Planned Start / End", "Baseline schedule dates."),
        ("Actual End", "Fill when Overall Status = Completed."),
        ("Dependency ID", "Feature ID that must complete first."),
        ("Remarks", "Blockers, decisions, links to PRD / ticket."),
    ]
    _cell(ws, 3, 1, "Field", bold=True, fill=HDR_FILL, color="FFFFFF")
    _cell(ws, 3, 2, "Definition", bold=True, fill=HDR_FILL, color="FFFFFF")
    for i, (a, b) in enumerate(rows, 4):
        _cell(ws, i, 1, a, bold=True)
        _cell(ws, i, 2, b, wrap=True)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 62


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    _project_cover(wb)
    _, last = _feature_register(wb)
    _dashboard(wb, last)
    _resource_plan(wb, last)
    _definitions(wb)
    wb.save(OUTPUT)
    print(f"Created {OUTPUT} — {len(_ROWS)} features, MNC standard register")


if __name__ == "__main__":
    main()
